"""FastAPI server for Sopal BIF Research.

Endpoints:
  POST /api/ask                         answer a question (streaming SSE)
  GET  /api/conversations               list recent conversations
  GET  /api/conversations/{id}          load one
  POST /api/conversations               create new (returns id)
  GET  /api/sources/{chunk_id}          fetch a source chunk by id
  GET  /api/health                      diagnostics

  GET  /                                serves the SPA shell
  GET  /static/...                      static assets

Run:
    OPENAI_API_KEY=sk-... uvicorn services.bif_research.api:app --port 8000
"""
from __future__ import annotations

import json
import logging
import os
import sqlite3
import sys
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from . import budget
from .indexer import CHUNKS_DB, BM25_PATH, CHROMA_PATH
from . import llm_config, name_index
from . import planner as planner_module
from .hard_pipeline import HardQuestionPipeline
from .pipeline import (
    FullPipeline,
    NaivePipeline,
    _format_history_for_answerer,
    _rewrite_with_history,
    route_question,
)


_TITLE_SYSTEM_PROMPT = (
    "Return a concise 3-6 word title for this Queensland construction-law "
    "research question. Title-case. No punctuation. No quotes. Output ONLY "
    "the title text — no preamble, no explanation."
)


def _generate_chat_title(question: str) -> str:
    """Generate a short title for a new conversation. Falls back to a
    truncated copy of the question if the LLM call fails. Cheap chain."""
    try:
        text, _ = llm_config.complete_chat(
            messages=[
                {"role": "system", "content": _TITLE_SYSTEM_PROMPT},
                {"role": "user", "content": question[:500]},
            ],
            kind="default",
            operation="chat-title",
            max_output_tokens=400,
        )
        title = (text or "").strip().strip('"').strip("'").strip()
        # Single line, no newlines
        title = title.split("\n")[0].strip()
        if 0 < len(title) <= 80:
            return title
    except Exception as e:
        log.info(f"chat-title generation failed: {e}")
    # Fallback: truncated question
    short = question.strip().split("\n")[0]
    if len(short) > 60:
        short = short[:57].rstrip() + "…"
    return short


def _resolved_case_citations(named_authorities: list[str]) -> list[dict]:
    """Resolve planner-named authorities to citation labels for the
    reading_cases status event. Each entry: {input, citation, case_id,
    case_name}. Misses get marked with case_id=None."""
    out: list[dict] = []
    seen_case_ids: set[str] = set()
    for name in named_authorities or []:
        m = name_index.lookup_case(name)
        if m and m.case_id not in seen_case_ids:
            seen_case_ids.add(m.case_id)
            out.append({
                "input": name,
                "case_id": m.case_id,
                "case_name": m.case_name,
                "citation": m.citation or m.case_name,
                "in_corpus": True,
            })
        elif not m:
            out.append({
                "input": name,
                "case_id": None,
                "case_name": "",
                "citation": name,
                "in_corpus": False,
            })
    return out

log = logging.getLogger("bif_research.api")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

ROOT = Path(__file__).resolve().parent
WEB = ROOT / "web"

# Conversations DB lives on the persistent disk in production. Override via
# BIF_CONV_DB env var. Falls back to the in-package store/ for local dev.
_default_conv_db = ROOT / "store" / "conversations.sqlite"
CONV_DB = Path(os.environ.get("BIF_CONV_DB", str(_default_conv_db)))

# Anonymous-user lifetime question cap (no sign-in required)
ANON_QUESTION_LIMIT = int(os.environ.get("BIF_ANON_LIMIT", "4"))
# Signed-in user per-UTC-day cap before requiring enterprise contact
SIGNED_DAILY_LIMIT = int(os.environ.get("BIF_SIGNED_DAILY_LIMIT", "30"))
# JWT secret + algorithm — must match the values server.py uses to sign
# `purchase_token` so we can verify it without a circular import. server.py
# reads `LEXIFILE_SECRET_KEY` and falls back to "dev-secret-key"; we mirror
# that exactly so signed users get recognised in production.
JWT_SECRET = (
    os.environ.get("LEXIFILE_SECRET_KEY")
    or os.environ.get("SECRET_KEY")
    or os.environ.get("JWT_SECRET")
    or "dev-secret-key"
)
JWT_ALG = os.environ.get("JWT_ALG", "HS256")

# User-document uploads
UPLOAD_DIR = Path(os.environ.get("BIF_UPLOAD_DIR", "/var/data/bif_uploads"))
MAX_UPLOAD_BYTES = int(os.environ.get("BIF_MAX_UPLOAD_BYTES", str(20 * 1024 * 1024)))   # 20 MB
MAX_DOC_TEXT_CHARS = int(os.environ.get("BIF_MAX_DOC_TEXT_CHARS", str(180_000)))         # ≈45k tokens
MAX_DOCS_PER_CONV = int(os.environ.get("BIF_MAX_DOCS_PER_CONV", "10"))


# ---------------------------------------------------------------------------
# Conversation store
# ---------------------------------------------------------------------------

def init_conv_db() -> sqlite3.Connection:
    CONV_DB.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(str(CONV_DB), check_same_thread=False)
    con.execute("""
        CREATE TABLE IF NOT EXISTS conversations (
            id TEXT PRIMARY KEY,
            title TEXT,
            created_at REAL,
            updated_at REAL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            conversation_id TEXT,
            role TEXT,
            content_json TEXT,
            created_at REAL,
            FOREIGN KEY (conversation_id) REFERENCES conversations(id)
        )
    """)
    con.execute("CREATE INDEX IF NOT EXISTS idx_msg_conv ON messages(conversation_id)")
    # Per-query cost ledger. One row per /api/ask call.
    # Not surfaced to the frontend; admin endpoints in /api/admin/costs* expose it.
    con.execute("""
        CREATE TABLE IF NOT EXISTS query_costs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            conversation_id TEXT,
            question TEXT,
            started_at REAL,
            finished_at REAL,
            elapsed_ms INTEGER,
            n_propositions INTEGER,
            n_sources INTEGER,
            confidence TEXT,
            refused INTEGER,
            cumulative_usd_before REAL,
            cumulative_usd_after REAL,
            total_cost_usd REAL,
            input_tokens INTEGER,
            output_tokens INTEGER,
            n_api_calls INTEGER,
            by_operation_json TEXT,
            calls_json TEXT
        )
    """)
    con.execute("CREATE INDEX IF NOT EXISTS idx_query_costs_conv ON query_costs(conversation_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_query_costs_started ON query_costs(started_at)")
    # Per-identifier query quota ledger (anon + signed). One row per asked
    # question. We compute lifetime/daily totals at check-time; trivially
    # cheap with the index.
    con.execute("""
        CREATE TABLE IF NOT EXISTS query_quota (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            identifier TEXT NOT NULL,   -- 'anon:<uuid>' or 'user:<email>'
            is_signed INTEGER NOT NULL,
            asked_at REAL NOT NULL,
            question TEXT
        )
    """)
    con.execute("CREATE INDEX IF NOT EXISTS idx_quota_id_time ON query_quota(identifier, asked_at)")
    # Migration: add identifier column to query_costs so the admin dashboard
    # can show who asked each query. Safe to run repeatedly — sqlite raises
    # OperationalError if the column already exists; we swallow that.
    try:
        con.execute("ALTER TABLE query_costs ADD COLUMN identifier TEXT")
    except sqlite3.OperationalError:
        pass
    # Same migration for conversations — owner identifier so /api/conversations
    # only returns chats the requesting user owns. Legacy rows (NULL) become
    # invisible to everyone.
    try:
        con.execute("ALTER TABLE conversations ADD COLUMN identifier TEXT")
    except sqlite3.OperationalError:
        pass
    con.execute("CREATE INDEX IF NOT EXISTS idx_conv_identifier ON conversations(identifier)")
    # Log of authorities the planner wanted to cite but the name_index
    # could not resolve — these are typically interstate/HCA cases not in
    # the Qld corpus, OR Qld cases that were never indexed (corpus gaps).
    # The dev portal aggregates this to surface coverage gaps.
    con.execute("""
        CREATE TABLE IF NOT EXISTS missing_authorities (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            seen_at         REAL NOT NULL,
            conversation_id TEXT,
            question        TEXT,
            authority       TEXT NOT NULL,
            identifier      TEXT
        )
    """)
    con.execute("CREATE INDEX IF NOT EXISTS idx_missing_auth_name ON missing_authorities(authority)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_missing_auth_time ON missing_authorities(seen_at)")
    # User-uploaded documents: a contract / payment claim / schedule that
    # the user attaches to a conversation. Extracted text is included in
    # the prompt context for every subsequent ask in the same conversation.
    con.execute("""
        CREATE TABLE IF NOT EXISTS uploaded_documents (
            id              TEXT PRIMARY KEY,
            conversation_id TEXT NOT NULL,
            filename        TEXT NOT NULL,
            mime            TEXT,
            size_bytes      INTEGER,
            n_chars         INTEGER,
            text            TEXT,
            uploaded_at     REAL,
            disk_path       TEXT
        )
    """)
    con.execute("CREATE INDEX IF NOT EXISTS idx_uploaded_conv ON uploaded_documents(conversation_id)")
    con.commit()
    return con


# ---------------------------------------------------------------------------
# Admin authentication
# ---------------------------------------------------------------------------

# Comma-separated list of emails allowed to hit /api/admin/* endpoints.
# Defaults to the build owner so the dashboard isn't open to the public.
ADMIN_EMAILS = {
    e.strip().lower()
    for e in os.environ.get("BIF_ADMIN_EMAILS", "edwardsheppard5@gmail.com").split(",")
    if e.strip()
}


def _require_admin(request: "Request") -> str:
    """Decode the Authorization JWT, confirm the email is in ADMIN_EMAILS,
    and return that email. Raises 401/403 otherwise."""
    email = _decode_user_email(request.headers.get("authorization"))
    if not email:
        raise HTTPException(401, "Sign in required")
    if email.lower() not in ADMIN_EMAILS:
        raise HTTPException(403, "Admin only")
    return email.lower()


_conv_con: sqlite3.Connection | None = None
_fast_pipeline: FullPipeline | None = None
_hard_pipeline: HardQuestionPipeline | None = None


def conv_con() -> sqlite3.Connection:
    global _conv_con
    if _conv_con is None:
        _conv_con = init_conv_db()
    return _conv_con


def get_pipelines() -> tuple[FullPipeline, HardQuestionPipeline]:
    """Lazy-init both pipelines. Routing happens in /api/ask after the
    planner runs once."""
    global _fast_pipeline, _hard_pipeline
    if _fast_pipeline is None:
        _fast_pipeline = FullPipeline(k_chunks=16, real_planner=True)
    if _hard_pipeline is None:
        _hard_pipeline = HardQuestionPipeline()
    return _fast_pipeline, _hard_pipeline


# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Warm the conversation DB
    conv_con()
    log.info("bif_research API ready")
    yield


app = FastAPI(lifespan=lifespan, title="Sopal BIF Research")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # tighten in production
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class AskRequest(BaseModel):
    question: str
    conversation_id: str | None = None
    pipeline_mode: str = "hybrid"   # "bm25" | "dense" | "hybrid"


class ConvCreateRequest(BaseModel):
    title: str | None = None


# ---------------------------------------------------------------------------
# Auth + per-identifier quota
# ---------------------------------------------------------------------------

def _decode_user_email(authorization_header: Optional[str]) -> Optional[str]:
    """Decode the JWT in an Authorization: Bearer <token> header. Returns
    the user's email if valid, else None. Mirrors server.py's purchase
    token format."""
    if not authorization_header:
        return None
    h = authorization_header.strip()
    if h.lower().startswith("bearer "):
        h = h[7:].strip()
    if not h:
        return None
    try:
        from jose import jwt as _jwt, JWTError
        payload = _jwt.decode(h, JWT_SECRET, algorithms=[JWT_ALG])
        # server.py stores email under 'sub' (standard) or 'email'
        return payload.get("sub") or payload.get("email")
    except Exception:
        return None


def _identify_request(authorization: Optional[str], anon_id: Optional[str]) -> tuple[str, bool, Optional[str]]:
    """Resolve who's asking. Returns (identifier, is_signed, email_or_none).

    Identifier format:
      'user:<email>'  if the JWT in Authorization decodes to an email
      'anon:<uuid>'   otherwise; uses X-Anon-ID header (client-generated)
    """
    email = _decode_user_email(authorization)
    if email:
        return f"user:{email.lower()}", True, email
    aid = (anon_id or "").strip()
    if not aid:
        # No anon-id supplied — generate one server-side. Client SHOULD send
        # X-Anon-ID for stable counting; absent that we can't track properly
        # so we still rate-limit but every request looks like a new user.
        # In practice the frontend always sends one.
        aid = uuid.uuid4().hex
    return f"anon:{aid}", False, None


def _quota_check(identifier: str, is_signed: bool) -> tuple[bool, dict]:
    """Return (allowed, info). info contains:
      - kind: 'anon' | 'signed'
      - used:  questions used so far (lifetime for anon, today UTC for signed)
      - limit: cap
      - remaining: max(0, limit-used)
    If not allowed, frontend should show the upgrade/contact gate.
    """
    con = conv_con()
    if is_signed:
        # Count today UTC
        now = datetime.now(timezone.utc)
        start_of_day = datetime(now.year, now.month, now.day, tzinfo=timezone.utc).timestamp()
        used = con.execute(
            "SELECT COUNT(*) FROM query_quota WHERE identifier=? AND asked_at>=?",
            (identifier, start_of_day),
        ).fetchone()[0]
        info = {"kind": "signed", "used": used, "limit": SIGNED_DAILY_LIMIT,
                "remaining": max(0, SIGNED_DAILY_LIMIT - used)}
        return used < SIGNED_DAILY_LIMIT, info
    used = con.execute(
        "SELECT COUNT(*) FROM query_quota WHERE identifier=?",
        (identifier,),
    ).fetchone()[0]
    info = {"kind": "anon", "used": used, "limit": ANON_QUESTION_LIMIT,
            "remaining": max(0, ANON_QUESTION_LIMIT - used)}
    return used < ANON_QUESTION_LIMIT, info


def _record_quota(identifier: str, is_signed: bool, question: str) -> None:
    con = conv_con()
    con.execute(
        "INSERT INTO query_quota (identifier, is_signed, asked_at, question) VALUES (?,?,?,?)",
        (identifier, 1 if is_signed else 0, time.time(), (question or "")[:500]),
    )
    con.commit()


def _gate_message(info: dict) -> str:
    if info.get("kind") == "anon":
        return ("You've used your free preview questions. Sign in (free) to "
                "get 30 questions per day.")
    return ("You've reached today's 30-question limit. For unlimited access, "
            "please contact info@sopal.com.au to arrange enterprise access.")


# ---------------------------------------------------------------------------
# Conversation endpoints
# ---------------------------------------------------------------------------

def _conv_owner(conv_id: str) -> Optional[str]:
    """Return the identifier that owns this conversation, or None if the
    row is missing OR is a legacy NULL-identifier row (treated as owned
    by no one)."""
    r = conv_con().execute(
        "SELECT identifier FROM conversations WHERE id = ?", (conv_id,)
    ).fetchone()
    return r[0] if r and r[0] else None


def _own_conv_or_404(conv_id: str, identifier: str) -> None:
    """Raise 404 if conv doesn't exist, 403 if it's owned by someone else."""
    owner = _conv_owner(conv_id)
    if owner is None:
        raise HTTPException(404, "Conversation not found")
    if owner != identifier:
        # 404, not 403, to avoid leaking existence
        raise HTTPException(404, "Conversation not found")


@app.get("/api/conversations")
def list_conversations(request: Request, limit: int = 30):
    identifier, _signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    rows = conv_con().execute(
        "SELECT id, title, created_at, updated_at FROM conversations "
        "WHERE identifier = ? "
        "ORDER BY updated_at DESC LIMIT ?",
        (identifier, limit),
    ).fetchall()
    return [{"id": r[0], "title": r[1], "created_at": r[2], "updated_at": r[3]} for r in rows]


@app.post("/api/conversations")
def create_conversation(payload: ConvCreateRequest, request: Request):
    identifier, _signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    conv_id = uuid.uuid4().hex[:12]
    now = time.time()
    title = payload.title or "New chat"
    conv_con().execute(
        "INSERT INTO conversations (id, title, created_at, updated_at, identifier) "
        "VALUES (?, ?, ?, ?, ?)",
        (conv_id, title, now, now, identifier),
    )
    conv_con().commit()
    return {"id": conv_id, "title": title}


@app.get("/api/conversations/{conv_id}")
def get_conversation(conv_id: str, request: Request):
    identifier, _signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    _own_conv_or_404(conv_id, identifier)
    conv = conv_con().execute(
        "SELECT id, title, created_at, updated_at FROM conversations WHERE id = ?", (conv_id,)
    ).fetchone()
    msgs = conv_con().execute(
        "SELECT role, content_json, created_at FROM messages "
        "WHERE conversation_id = ? ORDER BY id", (conv_id,)
    ).fetchall()
    return {
        "id": conv[0],
        "title": conv[1],
        "created_at": conv[2],
        "updated_at": conv[3],
        "messages": [
            {"role": r[0], "content": json.loads(r[1] or "{}"), "created_at": r[2]}
            for r in msgs
        ],
    }


@app.delete("/api/conversations/{conv_id}")
def delete_conversation(conv_id: str, request: Request):
    """Permanently delete a conversation and all its messages, documents,
    and uploaded files on disk. Owner-gated."""
    identifier, _signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    _own_conv_or_404(conv_id, identifier)
    # Best-effort: purge uploaded files on disk
    try:
        rows = conv_con().execute(
            "SELECT disk_path FROM uploaded_documents WHERE conversation_id=?",
            (conv_id,),
        ).fetchall()
        for r in rows:
            try: Path(r[0]).unlink(missing_ok=True)
            except Exception: pass
        conv_dir = UPLOAD_DIR / conv_id
        try: conv_dir.rmdir()
        except Exception: pass
    except Exception as e:
        log.warning(f"upload purge failed for conv {conv_id}: {e}")
    # DB deletes
    conv_con().execute("DELETE FROM uploaded_documents WHERE conversation_id=?", (conv_id,))
    conv_con().execute("DELETE FROM messages WHERE conversation_id=?", (conv_id,))
    conv_con().execute("DELETE FROM conversations WHERE id=?", (conv_id,))
    conv_con().commit()
    return {"ok": True, "deleted": conv_id}


# ---------------------------------------------------------------------------
# User-document uploads (per-conversation context)
# ---------------------------------------------------------------------------

def _extract_pdf_text(path: Path) -> str:
    """Try PyPDF2 first; fall back to pdfplumber which handles a wider
    range of PDF encodings. Returns empty string for scanned/image PDFs
    (those need OCR, which is a future enhancement)."""
    text = ""
    try:
        import PyPDF2
        out = []
        with open(path, "rb") as f:
            r = PyPDF2.PdfReader(f)
            for page in r.pages:
                try:
                    out.append(page.extract_text() or "")
                except Exception:
                    continue
        text = "\n".join(out).strip()
    except Exception as e:
        log.info(f"PyPDF2 extract failed for {path}: {e}")

    if len(text) < 200:
        try:
            import pdfplumber
            out = []
            with pdfplumber.open(str(path)) as pdf:
                for page in pdf.pages:
                    try:
                        out.append(page.extract_text() or "")
                    except Exception:
                        continue
            alt = "\n".join(out).strip()
            if len(alt) > len(text):
                text = alt
        except Exception as e:
            log.info(f"pdfplumber extract failed for {path}: {e}")

    return text


def _extract_docx_text(path: Path) -> str:
    try:
        import docx
    except Exception:
        return ""
    try:
        d = docx.Document(str(path))
        return "\n".join(p.text for p in d.paragraphs if p.text)
    except Exception as e:
        log.warning(f"docx extract failed for {path}: {e}")
        return ""


def _extract_xlsx_text(path: Path) -> str:
    try:
        import openpyxl
    except Exception:
        return ""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        out = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            out.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in cells):
                    out.append("\t".join(cells))
        wb.close()
        return "\n".join(out)
    except Exception as e:
        log.warning(f"xlsx extract failed for {path}: {e}")
        return ""


def _extract_image_ocr(path: Path) -> str:
    try:
        import pytesseract
        from PIL import Image
    except Exception:
        return ""
    try:
        img = Image.open(str(path))
        return pytesseract.image_to_string(img) or ""
    except Exception as e:
        log.warning(f"image OCR failed for {path}: {e}")
        return ""


def _extract_msg_text(path: Path) -> str:
    try:
        import extract_msg
    except Exception:
        return ""
    try:
        m = extract_msg.Message(str(path))
        parts = []
        if m.subject: parts.append(f"Subject: {m.subject}")
        if m.sender: parts.append(f"From: {m.sender}")
        if m.to: parts.append(f"To: {m.to}")
        if m.date: parts.append(f"Date: {m.date}")
        if m.body: parts.append("\n" + m.body)
        return "\n".join(parts)
    except Exception as e:
        log.warning(f"msg extract failed for {path}: {e}")
        return ""


def _extract_pdf_ocr(path: Path) -> str:
    """OCR-based fallback for image-only PDFs. Uses pdf2image (needs
    poppler-utils) → pytesseract per page. Slow but handles scanned
    decisions. Returns '' if dependencies are missing."""
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except Exception:
        return ""
    try:
        pages = convert_from_path(str(path), dpi=200)
    except Exception as e:
        log.warning(f"pdf2image failed for {path}: {e}")
        return ""
    out = []
    for i, img in enumerate(pages, 1):
        try:
            txt = pytesseract.image_to_string(img) or ""
            if txt.strip():
                out.append(f"--- page {i} ---\n{txt}")
        except Exception as e:
            log.info(f"OCR page {i} failed: {e}")
            continue
    return "\n\n".join(out)


def _extract_text_from_upload(path: Path, mime: str, filename: str) -> str:
    name = filename.lower()
    # PDF — text-based first; OCR as last-ditch fallback for scans
    if name.endswith(".pdf") or mime == "application/pdf":
        text = _extract_pdf_text(path)
        if len(text.strip()) < 200:
            ocr = _extract_pdf_ocr(path)
            if len(ocr) > len(text):
                text = ocr
        return text
    if name.endswith(".docx") or mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return _extract_docx_text(path)
    if name.endswith(".doc"):
        # Legacy Word — try docx2txt; if not installed, return empty.
        # textract / antiword are heavier system deps; skip for now.
        try:
            import docx2txt
            return (docx2txt.process(str(path)) or "").strip()
        except Exception as e:
            log.info(f"doc legacy extract not available for {path}: {e}")
            return ""
    if name.endswith((".xlsx", ".xlsm")) or mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    ):
        return _extract_xlsx_text(path)
    if name.endswith(".csv") or mime == "text/csv":
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return ""
    if name.endswith((".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp")) or mime.startswith("image/"):
        return _extract_image_ocr(path)
    if name.endswith(".msg"):
        return _extract_msg_text(path)
    if name.endswith(".eml"):
        try:
            from email import message_from_bytes
            from email.policy import default as email_policy
            msg = message_from_bytes(path.read_bytes(), policy=email_policy)
            body = msg.get_body(preferencelist=("plain", "html"))
            txt = ""
            if body:
                try: txt = body.get_content() or ""
                except Exception: txt = ""
            header = "\n".join(f"{k}: {v}" for k, v in msg.items())
            return header + "\n\n" + txt
        except Exception as e:
            log.warning(f"eml extract failed: {e}")
            return ""
    if name.endswith((".txt", ".md", ".rtf", ".log")) or mime.startswith("text/"):
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
            if name.endswith(".rtf"):
                # Best-effort RTF plain-text — strip control words crudely.
                import re
                raw = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", raw)
                raw = re.sub(r"[{}]", "", raw)
            return raw
        except Exception:
            return ""
    return ""


@app.post("/api/upload")
async def upload_document(
    request: Request,
    file: UploadFile = File(...),
    conversation_id: str = Form(...),
):
    """Accept a single document (PDF / DOCX / TXT / MD) up to 20 MB and
    attach it to a conversation. Subsequent /api/ask calls in the same
    conversation will include the extracted text in the prompt context."""
    # Quota gate (signed users + anon-id) — same identifier rules as /ask
    identifier, is_signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    # Conversation must exist AND be owned by the requester
    _own_conv_or_404(conversation_id, identifier)
    # Cap on docs per conversation
    n_existing = conv_con().execute(
        "SELECT COUNT(*) FROM uploaded_documents WHERE conversation_id=?",
        (conversation_id,),
    ).fetchone()[0]
    if n_existing >= MAX_DOCS_PER_CONV:
        raise HTTPException(409, f"Max {MAX_DOCS_PER_CONV} documents per conversation")

    # Read body and enforce size cap
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File exceeds {MAX_UPLOAD_BYTES // (1024*1024)} MB limit")
    if not raw:
        raise HTTPException(400, "Empty file")

    # Persist to disk (per-conv subdir on the persistent volume)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    conv_dir = UPLOAD_DIR / conversation_id
    conv_dir.mkdir(parents=True, exist_ok=True)
    doc_id = uuid.uuid4().hex[:16]
    safe_name = (file.filename or "upload.bin").replace("/", "_").replace("\\", "_")[:200]
    suffix = Path(safe_name).suffix or ""
    disk_path = conv_dir / f"{doc_id}{suffix}"
    disk_path.write_bytes(raw)

    # Extract text (best-effort, capped)
    text = _extract_text_from_upload(disk_path, file.content_type or "", safe_name)
    truncated = False
    if len(text) > MAX_DOC_TEXT_CHARS:
        text = text[:MAX_DOC_TEXT_CHARS]
        truncated = True

    conv_con().execute(
        """INSERT INTO uploaded_documents
           (id, conversation_id, filename, mime, size_bytes, n_chars, text, uploaded_at, disk_path)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            doc_id,
            conversation_id,
            safe_name,
            file.content_type or "",
            len(raw),
            len(text),
            text,
            time.time(),
            str(disk_path),
        ),
    )
    conv_con().commit()
    return {
        "id": doc_id,
        "filename": safe_name,
        "size_bytes": len(raw),
        "n_chars": len(text),
        "truncated": truncated,
        "extracted": bool(text.strip()),
    }


@app.get("/api/conversations/{conv_id}/documents")
def list_conversation_documents(conv_id: str, request: Request):
    identifier, _signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    _own_conv_or_404(conv_id, identifier)
    rows = conv_con().execute(
        "SELECT id, filename, mime, size_bytes, n_chars, uploaded_at "
        "FROM uploaded_documents WHERE conversation_id=? ORDER BY uploaded_at",
        (conv_id,),
    ).fetchall()
    return {
        "documents": [
            {
                "id": r[0], "filename": r[1], "mime": r[2],
                "size_bytes": r[3], "n_chars": r[4], "uploaded_at": r[5],
            }
            for r in rows
        ]
    }


@app.delete("/api/upload/{doc_id}")
def delete_document(doc_id: str, request: Request):
    identifier, _signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    row = conv_con().execute(
        "SELECT disk_path, conversation_id FROM uploaded_documents WHERE id=?",
        (doc_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "Not found")
    # Owner check via parent conversation
    _own_conv_or_404(row[1], identifier)
    try:
        Path(row[0]).unlink(missing_ok=True)
    except Exception as e:
        log.warning(f"failed to delete upload file: {e}")
    conv_con().execute("DELETE FROM uploaded_documents WHERE id=?", (doc_id,))
    conv_con().commit()
    return {"ok": True}


def _load_conv_documents_text(conv_id: str) -> list[dict]:
    """Return all uploaded documents for a conversation, ready to be
    prepended to the prompt context."""
    if not conv_id:
        return []
    rows = conv_con().execute(
        "SELECT id, filename, n_chars, text FROM uploaded_documents "
        "WHERE conversation_id=? ORDER BY uploaded_at",
        (conv_id,),
    ).fetchall()
    return [
        {"id": r[0], "filename": r[1], "n_chars": r[2], "text": r[3] or ""}
        for r in rows
    ]


# ---------------------------------------------------------------------------
# Source endpoint
# ---------------------------------------------------------------------------

@app.get("/api/sources/{chunk_id}")
def get_source(chunk_id: str):
    if not CHUNKS_DB.exists():
        raise HTTPException(503, "Index not built")
    con = sqlite3.connect(str(CHUNKS_DB), check_same_thread=False)
    con.row_factory = sqlite3.Row
    row = con.execute(
        "SELECT chunk_id, source_id, source_type, header, text, metadata_json "
        "FROM chunks WHERE chunk_id = ?", (chunk_id,)
    ).fetchone()
    con.close()
    if not row:
        raise HTTPException(404, "Chunk not found")
    return {
        "chunk_id": row["chunk_id"],
        "source_id": row["source_id"],
        "source_type": row["source_type"],
        "header": row["header"],
        "text": row["text"],
        "metadata": json.loads(row["metadata_json"] or "{}"),
    }


# ---------------------------------------------------------------------------
# Ask endpoint (SSE)
# ---------------------------------------------------------------------------

@app.get("/api/usage")
def get_usage(request: Request):
    """Return the caller's quota state. Anon callers should send X-Anon-ID."""
    identifier, is_signed, email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    allowed, info = _quota_check(identifier, is_signed)
    return {
        "is_signed": is_signed,
        "email": email,
        "allowed": allowed,
        **info,
    }


@app.post("/api/ask")
async def ask(payload: AskRequest, request: Request):
    """Stream the answer flow as SSE.

    Events emitted:
      status   - {"phase": "planning"|"retrieving"|"answering"|"verifying", "msg": str}
      answer   - the final dict from pipeline.answer(...)
      title    - {"conversation_id": str, "title": str}
      error    - {"error": str}
      done     - {}

    Authorization:
      Bearer JWT in the Authorization header for signed users (30/day cap).
      X-Anon-ID header (UUID) for anonymous users (4-question lifetime cap).
    """
    # Quota gate first — fail fast before spinning up the pipeline
    identifier, is_signed, _email = _identify_request(
        request.headers.get("authorization"),
        request.headers.get("x-anon-id"),
    )
    allowed, info = _quota_check(identifier, is_signed)
    if not allowed:
        return JSONResponse(
            status_code=429,
            content={
                "error": "quota_exceeded",
                "kind": info["kind"],
                "used": info["used"],
                "limit": info["limit"],
                "message": _gate_message(info),
            },
        )

    fast_pipe, hard_pipe = get_pipelines()

    # Persist user msg. Note: we DO NOT auto-set title from the question here.
    # The AI-generated title is set after the answer is composed and emitted
    # via SSE so the sidebar can slide it in.
    conv_id = payload.conversation_id
    is_first_turn = False
    if conv_id:
        # Owner check — block writing into someone else's conversation
        _own_conv_or_404(conv_id, identifier)
        # Check if this is the first turn so we know to generate a title
        existing_count = conv_con().execute(
            "SELECT COUNT(*) FROM messages WHERE conversation_id = ?",
            (conv_id,),
        ).fetchone()[0]
        is_first_turn = existing_count == 0
        conv_con().execute(
            "INSERT INTO messages (conversation_id, role, content_json, created_at) "
            "VALUES (?, ?, ?, ?)",
            (conv_id, "user", json.dumps({"text": payload.question}), time.time()),
        )
        conv_con().execute(
            "UPDATE conversations SET updated_at = ? WHERE id = ?",
            (time.time(), conv_id),
        )
        conv_con().commit()

    # Load conversation history (prior turns) so the pipeline can interpret
    # follow-up questions in context. Without this, "is it sufficient?" is
    # treated as a brand-new search and matches against any chunk containing
    # "sufficient" in the corpus.
    history = []
    if conv_id:
        rows = conv_con().execute(
            "SELECT role, content_json FROM messages "
            "WHERE conversation_id = ? ORDER BY id",
            (conv_id,),
        ).fetchall()
        for r in rows:
            try:
                content = json.loads(r[1] or "{}")
            except json.JSONDecodeError:
                content = {}
            # User messages store {"text": "..."}; assistant messages store the
            # full structured answer dict — strip to summary text downstream.
            if r[0] == "user":
                text = content.get("text", "")
                if text:
                    history.append({"role": "user", "content": text})
            else:
                history.append({"role": "assistant", "content": content})

    def gen():
        def emit(event: str, data: dict) -> bytes:
            return f"event: {event}\ndata: {json.dumps(data)}\n\n".encode()
        try:
            # Record this question against the caller's quota now (before any
            # work). If the pipeline crashes the user still spent a quota
            # slot — that's the right tradeoff vs giving free retries.
            try:
                _record_quota(identifier, is_signed, payload.question)
            except Exception as qerr:
                log.warning(f"failed to record quota: {qerr}")

            started_at = time.time()
            cum_before = budget.cumulative_usd()
            cap = budget.CostCapture()

            # NB: do NOT yield inside a `with cap:` block. Starlette runs
            # successive iterations of this sync generator in a thread pool;
            # the contextvar token created in __enter__ is invalidated when
            # the next iteration lands on a different thread. We keep each
            # `with cap:` block yield-free and let cap.calls accumulate
            # across multiple enter/exit cycles in different threads.

            yield emit("status", {"phase": "planning", "msg": "Planning"})

            # Load any documents the user attached to this conversation and
            # prepend their extracted text to the question. The planner +
            # answerer/reasoner all see this as part of the user turn so
            # they can read and reason over the document content.
            user_docs = _load_conv_documents_text(conv_id) if conv_id else []
            doc_blocks = []
            unreadable = []
            for d in user_docs:
                txt = (d.get("text") or "").strip()
                if not txt:
                    unreadable.append(d.get("filename", "(unnamed)"))
                    continue
                doc_blocks.append(
                    f"=== ATTACHED DOCUMENT: {d['filename']} ===\n{txt}"
                )
            preamble_parts = []
            if doc_blocks:
                preamble_parts.append(
                    "The user has attached the following document(s) to this "
                    "conversation. Read them and use them as part of the factual "
                    "context for your answer. They are NOT primary law and must "
                    "not be cited as authority.\n\n"
                    + "\n\n".join(doc_blocks)
                    + "\n\n=== END OF ATTACHED DOCUMENTS ==="
                )
            if unreadable:
                preamble_parts.append(
                    "NOTE: the user attached the following document(s) but no "
                    "text could be extracted (likely a scanned/image-only PDF "
                    "that requires OCR). Tell the user explicitly that you "
                    "could not read the file and suggest they paste the text or "
                    "upload a text-based PDF/DOCX:\n  - "
                    + "\n  - ".join(unreadable)
                )
            if preamble_parts:
                question_with_docs = (
                    "\n\n".join(preamble_parts)
                    + "\n\nUSER QUESTION:\n" + payload.question
                )
            else:
                question_with_docs = payload.question

            # Phase A — planner + routing (no yields inside)
            with cap:
                rewritten = _rewrite_with_history(question_with_docs, history)
                plan = planner_module.plan(rewritten, real=True)
                route = route_question(plan, rewritten)

            yield emit("status", {"phase": "routing",
                                  "msg": f"Routed to {route} path"})
            if route == "hard":
                resolved_cases = _resolved_case_citations(plan.named_authorities or [])
                in_corpus = [c for c in resolved_cases if c["in_corpus"]]
                missed = [c for c in resolved_cases if not c["in_corpus"]]
                yield emit("status", {
                    "phase": "reading_cases",
                    "msg": "Reading judgments",
                    "cases": [c["citation"] for c in in_corpus],
                    "missed": [c["citation"] for c in missed],
                })
                # Log the planner-wanted-but-not-indexed authorities so the
                # dev portal can surface coverage gaps over time.
                if missed:
                    try:
                        rows = [
                            (time.time(), conv_id, payload.question, c["citation"], identifier)
                            for c in missed
                        ]
                        conv_con().executemany(
                            "INSERT INTO missing_authorities "
                            "(seen_at, conversation_id, question, authority, identifier) "
                            "VALUES (?, ?, ?, ?, ?)",
                            rows,
                        )
                        conv_con().commit()
                    except Exception as merr:
                        log.warning(f"failed to log missing_authorities: {merr}")
            else:
                yield emit("status", {"phase": "retrieving",
                                      "msg": "Retrieving sources…"})

            # Phase B — answer composition + title generation (no yields inside)
            generated_title: str | None = None
            with cap:
                if route == "hard":
                    answer = hard_pipe.answer_with_plan(
                        rewritten, plan, history=history,
                        original_question=payload.question,
                    )
                else:
                    answer = fast_pipe.answer_with_plan(
                        rewritten, plan, history=history,
                        original_question=payload.question,
                    )
                # Generate a chat title for new conversations
                if conv_id and is_first_turn:
                    generated_title = _generate_chat_title(payload.question)
            cum_after = budget.cumulative_usd()
            elapsed = time.time() - started_at
            answer["_elapsed_ms"] = int(elapsed * 1000)
            answer["_route"] = route

            # Persist + emit the title before the answer so the sidebar
            # updates while the user is reading
            if generated_title and conv_id:
                try:
                    conv_con().execute(
                        "UPDATE conversations SET title = ? WHERE id = ?",
                        (generated_title, conv_id),
                    )
                    conv_con().commit()
                except Exception as title_err:
                    log.warning(f"failed to persist chat title: {title_err}")
                yield emit("title", {
                    "conversation_id": conv_id,
                    "title": generated_title,
                })

            yield emit("status", {"phase": "verifying", "msg": "Verifying citations…"})
            yield emit("answer", answer)
            cap_summary = cap.summary()
            # Persist per-query cost row (server-side only, not in answer payload)
            try:
                conv_con().execute(
                    """INSERT INTO query_costs
                       (conversation_id, question, started_at, finished_at, elapsed_ms,
                        n_propositions, n_sources, confidence, refused,
                        cumulative_usd_before, cumulative_usd_after, total_cost_usd,
                        input_tokens, output_tokens, n_api_calls,
                        by_operation_json, calls_json, identifier)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        conv_id,
                        payload.question,
                        started_at,
                        time.time(),
                        int(elapsed * 1000),
                        len(answer.get("propositions") or []),
                        len(answer.get("sources") or []),
                        answer.get("confidence", ""),
                        1 if answer.get("refused") else 0,
                        cum_before,
                        cum_after,
                        cap_summary["total_usd"],
                        cap_summary["input_tokens"],
                        cap_summary["output_tokens"],
                        cap_summary["n_api_calls"],
                        json.dumps(cap_summary["by_operation"]),
                        json.dumps(cap_summary["calls"]),
                        identifier,
                    ),
                )
                conv_con().commit()
                log.info(
                    f"query cost: ${cap_summary['total_usd']:.4f} "
                    f"({cap_summary['n_api_calls']} api calls, "
                    f"{cap_summary['input_tokens']}/{cap_summary['output_tokens']} tok, "
                    f"{int(elapsed*1000)}ms) — '{payload.question[:60]}'"
                )
            except Exception as cost_err:
                log.warning(f"failed to persist query_costs row: {cost_err}")

            if conv_id:
                conv_con().execute(
                    "INSERT INTO messages (conversation_id, role, content_json, created_at) "
                    "VALUES (?, ?, ?, ?)",
                    (conv_id, "assistant", json.dumps(answer), time.time()),
                )
                conv_con().execute(
                    "UPDATE conversations SET updated_at = ? WHERE id = ?",
                    (time.time(), conv_id),
                )
                conv_con().commit()
            yield emit("done", {})
        except budget.BudgetExceeded as e:
            yield emit("error", {"error": "Budget cap reached.", "detail": str(e)})
        except Exception as e:
            log.exception("ask failed")
            yield emit("error", {"error": str(e)})

    return StreamingResponse(gen(), media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Admin: per-query cost endpoints (NOT linked from the frontend).
# Designed to slot into the Sopal admin console later. No auth here yet —
# expose only on localhost. When integrated into the Sopal site, wrap
# behind the existing /admin/page-style auth gate.
# ---------------------------------------------------------------------------

@app.get("/api/admin/costs")
def admin_costs(request: Request, limit: int = 100, conversation_id: str | None = None):
    """Recent per-query cost rows. Most recent first."""
    _require_admin(request)
    where = ""
    params: list = []
    if conversation_id:
        where = "WHERE conversation_id = ?"
        params.append(conversation_id)
    params.append(limit)
    rows = conv_con().execute(
        f"""SELECT id, conversation_id, question, started_at, finished_at, elapsed_ms,
                   n_propositions, n_sources, confidence, refused,
                   total_cost_usd, input_tokens, output_tokens, n_api_calls,
                   by_operation_json, identifier
            FROM query_costs
            {where}
            ORDER BY started_at DESC
            LIMIT ?""",
        params,
    ).fetchall()
    out = []
    for r in rows:
        out.append({
            "id": r[0], "conversation_id": r[1], "question": r[2],
            "started_at": r[3], "finished_at": r[4], "elapsed_ms": r[5],
            "n_propositions": r[6], "n_sources": r[7],
            "confidence": r[8], "refused": bool(r[9]),
            "total_cost_usd": r[10],
            "input_tokens": r[11], "output_tokens": r[12], "n_api_calls": r[13],
            "by_operation": json.loads(r[14] or "{}"),
            "identifier": r[15] or "",
        })
    return {"queries": out, "n": len(out)}


@app.get("/api/admin/queries")
def admin_queries(request: Request, limit: int = 200):
    """Dev-portal feed: every recent query joined with the conversation
    title, formatted for direct table display. Admin only."""
    _require_admin(request)
    rows = conv_con().execute(
        """SELECT qc.id, qc.conversation_id, qc.question, qc.started_at,
                  qc.elapsed_ms, qc.confidence, qc.refused,
                  qc.total_cost_usd, qc.input_tokens, qc.output_tokens,
                  qc.n_api_calls, qc.identifier, qc.by_operation_json,
                  qc.n_propositions, qc.n_sources,
                  c.title
           FROM query_costs qc
           LEFT JOIN conversations c ON c.id = qc.conversation_id
           ORDER BY qc.started_at DESC
           LIMIT ?""",
        (limit,),
    ).fetchall()
    out = []
    for r in rows:
        ident = r[11] or ""
        is_signed = ident.startswith("user:")
        display_who = ident[5:] if is_signed else ("anon " + ident[5:13] if ident.startswith("anon:") else ident)
        try: by_op = json.loads(r[12] or "{}")
        except: by_op = {}
        # Identify route from operation names: hard pipeline uses 'reasoner'
        route = "hard" if any(k.startswith("reasoner") or k.startswith("read-case") for k in by_op) else "fast"
        out.append({
            "id": r[0],
            "conversation_id": r[1],
            "question": r[2],
            "started_at": r[3],
            "elapsed_ms": r[4],
            "confidence": r[5] or "",
            "refused": bool(r[6]),
            "total_cost_usd": r[7] or 0,
            "input_tokens": r[8] or 0,
            "output_tokens": r[9] or 0,
            "n_api_calls": r[10] or 0,
            "identifier": ident,
            "is_signed": is_signed,
            "who": display_who,
            "route": route,
            "n_propositions": r[13] or 0,
            "n_sources": r[14] or 0,
            "conv_title": r[15] or "",
        })
    return {"queries": out, "n": len(out)}


@app.get("/api/admin/missing-authorities")
def admin_missing_authorities(request: Request, days: int = 90, limit: int = 500):
    """Aggregate of authorities the planner named but the corpus didn't
    have. Returns a per-authority roll-up plus the most-recent N raw rows
    so we can drill down."""
    _require_admin(request)
    since = time.time() - days * 86400
    summary = conv_con().execute(
        """SELECT authority,
                  COUNT(*)            AS n_misses,
                  MAX(seen_at)        AS last_seen,
                  MIN(seen_at)        AS first_seen,
                  COUNT(DISTINCT conversation_id) AS n_conversations,
                  COUNT(DISTINCT identifier)      AS n_distinct_users
           FROM missing_authorities
           WHERE seen_at >= ?
           GROUP BY authority
           ORDER BY n_misses DESC, last_seen DESC""",
        (since,),
    ).fetchall()
    recent = conv_con().execute(
        """SELECT id, seen_at, conversation_id, question, authority, identifier
           FROM missing_authorities
           WHERE seen_at >= ?
           ORDER BY seen_at DESC
           LIMIT ?""",
        (since, limit),
    ).fetchall()
    return {
        "days": days,
        "summary": [
            {
                "authority": r[0],
                "n_misses": r[1],
                "last_seen": r[2],
                "first_seen": r[3],
                "n_conversations": r[4],
                "n_distinct_users": r[5],
            }
            for r in summary
        ],
        "recent": [
            {
                "id": r[0],
                "seen_at": r[1],
                "conversation_id": r[2],
                "question": r[3],
                "authority": r[4],
                "identifier": r[5] or "",
            }
            for r in recent
        ],
    }


@app.get("/api/admin/users")
def admin_users(request: Request, days: int = 30):
    """Per-user query counts + cost over the last N days (default 30)."""
    _require_admin(request)
    since = time.time() - days * 86400
    rows = conv_con().execute(
        """SELECT identifier,
                  COUNT(*)               AS n_queries,
                  COALESCE(SUM(total_cost_usd),0) AS cost,
                  COALESCE(SUM(input_tokens),0)   AS in_tok,
                  COALESCE(SUM(output_tokens),0)  AS out_tok,
                  MIN(started_at)        AS first_seen,
                  MAX(started_at)        AS last_seen
           FROM query_costs
           WHERE identifier IS NOT NULL AND started_at >= ?
           GROUP BY identifier
           ORDER BY last_seen DESC""",
        (since,),
    ).fetchall()
    out = []
    for r in rows:
        ident = r[0] or ""
        is_signed = ident.startswith("user:")
        out.append({
            "identifier": ident,
            "is_signed": is_signed,
            "who": ident[5:] if is_signed else ("anon " + ident[5:13] if ident.startswith("anon:") else ident),
            "n_queries": r[1],
            "total_cost_usd": round(r[2], 6),
            "input_tokens": r[3],
            "output_tokens": r[4],
            "first_seen": r[5],
            "last_seen": r[6],
        })
    return {"users": out, "days": days, "n": len(out)}


@app.get("/api/admin/costs/summary")
def admin_costs_summary(request: Request):
    """Aggregate cost stats — total, today, last 24h, average per query."""
    _require_admin(request)
    now = time.time()
    day_ago = now - 24 * 3600
    week_ago = now - 7 * 24 * 3600
    midnight = now - (now % 86400)  # crude UTC midnight
    rows = conv_con().execute(
        "SELECT COUNT(*), COALESCE(SUM(total_cost_usd), 0), "
        "COALESCE(AVG(total_cost_usd), 0), COALESCE(AVG(elapsed_ms), 0), "
        "COALESCE(SUM(input_tokens), 0), COALESCE(SUM(output_tokens), 0) "
        "FROM query_costs"
    ).fetchone()
    today = conv_con().execute(
        "SELECT COUNT(*), COALESCE(SUM(total_cost_usd), 0) FROM query_costs WHERE started_at >= ?",
        (midnight,),
    ).fetchone()
    last24 = conv_con().execute(
        "SELECT COUNT(*), COALESCE(SUM(total_cost_usd), 0) FROM query_costs WHERE started_at >= ?",
        (day_ago,),
    ).fetchone()
    last7 = conv_con().execute(
        "SELECT COUNT(*), COALESCE(SUM(total_cost_usd), 0) FROM query_costs WHERE started_at >= ?",
        (week_ago,),
    ).fetchone()
    return {
        "all_time": {
            "n_queries": rows[0],
            "total_cost_usd": round(rows[1], 6),
            "avg_cost_per_query_usd": round(rows[2], 6),
            "avg_elapsed_ms": round(rows[3], 0),
            "input_tokens": rows[4],
            "output_tokens": rows[5],
        },
        "today_utc": {"n_queries": today[0], "cost_usd": round(today[1], 6)},
        "last_24h": {"n_queries": last24[0], "cost_usd": round(last24[1], 6)},
        "last_7d": {"n_queries": last7[0], "cost_usd": round(last7[1], 6)},
        "cumulative_spend_usd": budget.cumulative_usd(),
        "remaining_usd": budget.remaining_usd(),
    }


@app.get("/api/admin/costs/{cost_id}")
def admin_cost_detail(cost_id: int, request: Request):
    """Full detail of one cost row, including per-call breakdown."""
    _require_admin(request)
    r = conv_con().execute(
        "SELECT id, conversation_id, question, started_at, finished_at, elapsed_ms, "
        "n_propositions, n_sources, confidence, refused, "
        "cumulative_usd_before, cumulative_usd_after, total_cost_usd, "
        "input_tokens, output_tokens, n_api_calls, "
        "by_operation_json, calls_json "
        "FROM query_costs WHERE id = ?",
        (cost_id,),
    ).fetchone()
    if not r:
        raise HTTPException(404, "Not found")
    return {
        "id": r[0], "conversation_id": r[1], "question": r[2],
        "started_at": r[3], "finished_at": r[4], "elapsed_ms": r[5],
        "n_propositions": r[6], "n_sources": r[7],
        "confidence": r[8], "refused": bool(r[9]),
        "cumulative_usd_before": r[10], "cumulative_usd_after": r[11],
        "total_cost_usd": r[12],
        "input_tokens": r[13], "output_tokens": r[14], "n_api_calls": r[15],
        "by_operation": json.loads(r[16] or "{}"),
        "calls": json.loads(r[17] or "[]"),
    }


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    info = {
        "ok": True,
        "cumulative_spend_usd": budget.cumulative_usd(),
        "remaining_usd": budget.remaining_usd(),
        "chunks_db_present": CHUNKS_DB.exists(),
        "bm25_present": BM25_PATH.exists(),
        "chroma_present": (CHROMA_PATH / "chroma.sqlite3").exists(),
    }
    if CHUNKS_DB.exists():
        con = sqlite3.connect(str(CHUNKS_DB), check_same_thread=False)
        n = con.execute("SELECT COUNT(*) FROM chunks").fetchone()[0]
        by_type = {
            r[0]: r[1]
            for r in con.execute(
                "SELECT source_type, COUNT(*) FROM chunks GROUP BY source_type"
            ).fetchall()
        }
        info["chunks_total"] = n
        info["chunks_by_type"] = by_type
        con.close()
    return info


# ---------------------------------------------------------------------------
# Static frontend
# ---------------------------------------------------------------------------

if WEB.exists():
    # /static/* -> any file under web/ (styles.css, app.js, future assets).
    # When the app is mounted at /ai in production, these become
    # /ai/static/styles.css etc — matching the absolute paths in
    # index.html. Local-dev clients hit /static/styles.css with no prefix.
    app.mount("/static", StaticFiles(directory=str(WEB)), name="static")

    @app.get("/")
    def index():
        return FileResponse(str(WEB / "index.html"))

    @app.get("/admin")
    def admin_dashboard():
        return FileResponse(str(WEB / "admin.html"))
